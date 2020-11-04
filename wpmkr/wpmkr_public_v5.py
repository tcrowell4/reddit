# pandas version 1.1.3 required
# xlrd version 1.2.0 required
# xlwt required
# xlsxwriter version 1.3.7 required
# openpyxl version 3.0.4 required

# https://github.com/py3nub/wpmkr
# author: reddit user u/pfnub

# importing required modules and libraries
"""
Version 5 will be to:
 remove redundant iters through dfcat
 eliminate copying the template file:
    we should be able to open the template file with openpyxl and just write
    the new file
 reformatted code using 'black'   https://github.com/psf/black
"""

import os
import pandas as pd
import shutil
import openpyxl
from openpyxl.utils import get_column_letter
import datetime


def process_workplans():
    # setting the main directory path to work in
    # main_path = "mypath
    main_path = "mypath"

    # print("Debug:  cwd",os.getcwd())
    wp_data_file = os.path.join(main_path, "wpdata.xlsx")
    wp_template = os.path.join(main_path, "Work Plan Template v2.xlsx")
    wp_folder = os.path.join(main_path, "Workplan_Folder")

    if not os.path.exists(wp_folder):
        os.mkdir(wp_folder)

    # creating dataframe for the Activity Data tab in the data sheet
    AD_cols = [
        "Biditem",
        "Bid Desc",
        "Description",
        "Quan",
        "Unit",
        "Shifts",
        "MH",
        "Total Labor",
    ]
    dfAD = pd.read_excel(wp_data_file, sheet_name="Activity Data", usecols=AD_cols)

    # replacing unwanted characters in the column names to be referenced later
    # print("Debug:  dfAD.columns\n",dfAD.columns)
    dfAD.columns = [c.replace(" ", "_") for c in dfAD.columns]
    dfAD.columns = [c.replace(".", "") for c in dfAD.columns]
    dfAD.columns = [c.replace("/", "") for c in dfAD.columns]
    # print("Debug:  POST replace dfAD.columns\n",dfAD.columns)

    # print("Debug:  dfAD\n",dfAD)
    # adding a prefix to the columns to know which dataframe we are pulling from
    dfAD = dfAD.add_prefix("AD_")
    # print("Debug:  dfAD add prefix\n",dfAD)

    # maybe don't need this AD_cols = list(dfAD.columns)
    # setting constraints on what we want to be used from the dataframe
    dfAD = dfAD[dfAD.AD_MH >= 500]
    dfAD = dfAD[dfAD.AD_Biditem <= 9000]
    # resetting the index to make the dataframe easier to read
    dfAD = dfAD.reset_index(drop=True)
    dfAD.index += 1

    # creating lists of certain columns to set more constraints in the next dataframe
    AD_biditem_list = sorted(set(dfAD["AD_Biditem"].tolist()))
    AD_bid_desc_list = sorted(set(dfAD["AD_Bid_Desc"].tolist()))
    AD_description_list = sorted(set(dfAD["AD_Description"].tolist()))
    # print("Debug:  AD Lists\n",AD_biditem_list, AD_bid_desc_list, AD_description_list)

    # creating dataframe for the Resource Data tab in the data sheet
    RD_cols = [
        "Biditem",
        "Bid Desc.",
        "Actv Desc.",
        "Description",
        "Quantity",
        "Unit",
        "Unit Cost",
        "Pcs/Wste",
        "Total",
    ]

    dfRD = pd.read_excel(wp_data_file, sheet_name="Resource Data", usecols=RD_cols)
    # print("Debug:  dfRD\n",dfRD)

    # replacing unusable characters in the column names to be referenced later
    dfRD.columns = [c.replace(" ", "_") for c in dfRD.columns]
    dfRD.columns = [c.replace(".", "") for c in dfRD.columns]
    dfRD.columns = [c.replace("/", "") for c in dfRD.columns]
    # adding a prefix to the columns to know which dataframe we are pulling from
    dfRD = dfRD.add_prefix("RD_")
    # maybe don't need this RD_cols = list(dfRD.columns)
    # print("Debug:  Post replace/prefix dfRD\n",dfRD)

    # setting constraints on what we want to be used from the dataframe
    dfRD = dfRD[(dfRD.RD_Unit == "MH") | (dfRD.RD_Unit == "HR")]
    dfRD = dfRD[dfRD.RD_Biditem <= 9000]
    # print("Debug:  Post constraints dfRD\n",dfRD)

    # comparing to dfAD to remove rows that do not match
    # print("Debug: RD_Biditem\n", dfRD['RD_Biditem'], dfRD['RD_Biditem'].isin(AD_biditem_list))
    # print("Debug: RD_Bid_Desc\n", dfRD['RD_Bid_Desc'], dfRD['RD_Bid_Desc'].isin(AD_bid_desc_list))
    # print("Debug: RD_Actv_Desc\n", dfRD['RD_Actv_Desc'], dfRD['RD_Actv_Desc'].isin(AD_description_list))

    dfRD = dfRD[dfRD["RD_Biditem"].isin(AD_biditem_list)]
    dfRD = dfRD[dfRD["RD_Bid_Desc"].isin(AD_bid_desc_list)]
    dfRD = dfRD[dfRD["RD_Actv_Desc"].isin(AD_description_list)]
    print("Debug:  Post remove rows dfRD\n", dfRD)

    # resetting the index to make the dataframe easier to read
    dfRD = dfRD.reset_index(drop=True)
    dfRD.index += 1

    # creating lists of the matching rows in dfAD and dfRD to reset the index
    # for dfRD to match dfAD and only contain
    # rows we want
    tuples_list = []
    index_list = []
    for rowAD in dfAD.itertuples():
        for rowRD in dfRD.itertuples():
            # print("Debug: Biditem ",rowAD.AD_Biditem,rowRD.RD_Biditem)
            # print("Debug: Bid_Desc ",rowAD.AD_Bid_Desc, rowRD.RD_Bid_Desc)
            # print("Debug: Description ",rowAD.AD_Description,rowRD.RD_Actv_Desc)
            if (
                rowAD.AD_Biditem == rowRD.RD_Biditem
                and rowAD.AD_Bid_Desc == rowRD.RD_Bid_Desc
                and rowAD.AD_Description == rowRD.RD_Actv_Desc
            ):
                tuples_list.append(rowRD)
                index_list.append(rowAD.Index)
    # print("Debug: tuples_list index_list",tuples_list,index_list)

    # creating dfRD2 that has the matching index and rows we want
    dfRD2 = pd.DataFrame(
        tuples_list,
        columns=[
            "RD_Index",
            "RD_Biditem",
            "RD_Bid_Desc",
            "RD_Actv_Desc",
            "RD_Description",
            "RD_Quantity",
            "RD_Unit",
            "RD_Unit_Cost",
            "RD_PcsWste",
            "RD_Total",
        ],
    )
    # replacing unusable characters in the column names to be referenced later
    # print("Debug:  Initial dfRD2\n",dfRD2)
    dfRD2.columns = [c.replace(" ", "_") for c in dfRD2.columns]
    dfRD2.columns = [c.replace(".", "") for c in dfRD2.columns]
    dfRD2.columns = [c.replace("/", "") for c in dfRD2.columns]
    # resetting the index to match dfAD
    dfRD2.index = index_list
    # dropping the old index
    dfRD2.drop(["RD_Index"], axis=1, inplace=True)
    # print("Debug:  Interim dfRD2\n",dfRD2)

    # removing special characters from the dataframes so that the descriptions can be used as filenames and folders later
    spec_chars = ["<", ">", ":", '"', "/", "\\", "|", "*"]
    for char in spec_chars:
        dfAD["AD_Description"] = dfAD["AD_Description"].str.replace(char, "")
    for char in spec_chars:
        dfRD2["RD_Actv_Desc"] = dfRD2["RD_Actv_Desc"].str.replace(char, "")

    print("Debug:  POST dfRD2\n", dfRD2)
    print("Debug:  dfAD\n", dfAD)
    # combining dfAD and dfRD2 so they are in one place
    dfcat = pd.concat([dfAD, dfRD2], axis=1, join="inner")
    print("Debug:  dfcat\n", dfcat)

    completed_by = "Python"
    job_number = "123"
    last_bidkey = ""

    for rowcat in dfcat.itertuples():
        # print("Debug: rowcat\n", rowcat)
        bid_num = str(int(rowcat.AD_Biditem))
        act_desc = str(rowcat.AD_Description)
        bidkey = bid_num + "_" + act_desc

        # iterating through each row of dfcat to place the information where it goes in the new workplan files
        # This currently opens a workbook and updates  it for the current rows
        # and then writes it ...
        # Then reloads the same workbook if there is more information on the next row
        # and depends on whether something has been written on a line to cause it to go to the
        # next lines in the sheet for  RD  data
        if bidkey != last_bidkey:
            # saving each iteration to it's respective work plan
            # write work book if it isn't the first time
            if last_bidkey != "":
                # saving each iteration to it's respective work plan
                wb.save(wb_name)
                print(
                    wb_name
                    + " has been updated. When this message changes, the workplan is complete."
                )

            # update last_bidkey
            last_bidkey = bidkey

            # These are the starting positions for writing the RD data
            # Used to avoid creating the boundary tables
            row_MH = 34
            col_MH = 11
            row_HR = 58
            col_HR = 11

            # initialize for change of bidkey
            temp_dir = os.path.join(wp_folder, bidkey)
            if not os.path.exists(temp_dir):
                os.mkdir(temp_dir)

            wb_name = os.path.join(temp_dir, bidkey + ".xlsx")
            # print("Debug: wb_name\n", wb_name, bidkey, temp_dir)
            # open the template directly and then write out the new file
            wb = openpyxl.load_workbook(wp_template)
            ws = wb["WORKPLAN"]
            # print("Debug: wb\n", wb)
            # print("Debug: ws\n", ws)

            # placing the budget information into the workplan
            # print("Debug: ws['C3'].value", ws['C3'].value)
            # no need to check if empty ... only written once
            ws["C3"] = completed_by
            ws["B4"] = job_number
            ws["K12"] = rowcat.AD_Description
            ws["M12"] = bid_num
            ws["N12"] = rowcat.AD_Quan
            ws["O12"] = rowcat.AD_Unit
            ws["R12"] = rowcat.AD_MH
            ws["T12"] = rowcat.AD_Shifts
            ws["U12"] = rowcat.AD_Total_Labor

        # deciding where the information goes depending if it's man hours or equipment hours
        if rowcat.RD_Unit == "MH":
            _ = ws.cell(column=col_MH, row=row_MH, value=rowcat.RD_Description)
            _ = ws.cell(column=col_MH + 1, row=row_MH, value=rowcat.RD_Unit_Cost)
            _ = ws.cell(
                column=col_MH + 2,
                row=row_MH,
                value=round(rowcat.RD_Quantity / rowcat.AD_Shifts, 2),
            )
            _ = ws.cell(column=col_MH + 3, row=row_MH, value=round(rowcat.RD_Total))
            row_MH += 1
        else:
            _ = ws.cell(column=col_HR, row=row_HR, value=rowcat.RD_Description)
            _ = ws.cell(
                column=col_HR + 1,
                row=row_HR,
                value=round(rowcat.RD_Quantity / rowcat.AD_Shifts, 2),
            )
            _ = ws.cell(column=col_HR + 2, row=row_HR, value=rowcat.RD_Total)
            row_HR += 1

    # saving each iteration to it's respective work plan
    wb.save(wb_name)  # write the final final
    # shows you in the command prompt what is being worked on
    print(
        wb_name
        + " has been updated. When this message changes, the workplan is complete."
    )

    """
    these are helper excel files that can be created to view each dataframe in excel if needed
    dfcat = pd.concat([dfAD,dfRD2], axis=1, join='inner')
    with pd.ExcelWriter('wpdata2.xlsx') as writer:
        dfcat.to_excel(writer, sheet_name='Concat', index=False)
        dfAD.to_excel(writer, sheet_name='AD', index=False)
        dfRD.to_excel(writer, sheet_name='RD', index=False)
        dfRD2.to_excel(writer, sheet_name='RD2', index=False)

    with pd.ExcelWriter('debug.xlsx') as writer:
        dfcat.to_excel(writer, sheet_name='debug', index=False)
    """


def main():

    begin_time = datetime.datetime.now()

    """
    Input Files:
    Work Plan source data:      mypath/wpdata.xlsx
    Work Plan base Template:    mypath/Work Plan Template v2

    Output:
    Work PLAN Folder location:  mypath/Workplan_Folder
    """
    process_workplans()

    print(
        "All work plans have been completed successfully\nTime to complete all work plans: "
        + str(datetime.datetime.now() - begin_time)
    )
    input("Press Enter to Escape")


if __name__ == "__main__":
    main()
